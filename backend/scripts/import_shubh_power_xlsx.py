from __future__ import annotations

import argparse
import hashlib
import re
from datetime import UTC, datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from pymongo import UpdateOne

from backend.app.core.config import get_settings
from backend.scripts.common import run


CITY_CENTERS: dict[str, tuple[float, float]] = {
    "delhi": (28.6139, 77.2090),
    "new delhi": (28.6139, 77.2090),
    "noida": (28.5355, 77.3910),
    "greater noida": (28.4744, 77.5040),
    "gurugram": (28.4595, 77.0266),
    "gurgaon": (28.4595, 77.0266),
    "ghaziabad": (28.6692, 77.4538),
    "faridabad": (28.4089, 77.3178),
}

REGION_HINTS = {
    "west": "Delhi",
    "south": "New Delhi",
    "north": "Delhi",
    "east": "Delhi",
    "central": "New Delhi",
    "gurugram": "Gurugram",
    "gurgaon": "Gurugram",
    "noida": "Noida",
    "ghaziabad": "Ghaziabad",
    "faridabad": "Faridabad",
}


def cell_text(value) -> str:
    return "" if value is None else str(value).strip()


def normalized_key(value: str) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return key[:96] or uuid4().hex[:12]


def stable_station_id(*parts: str) -> str:
    raw = "|".join(part for part in parts if part)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10]
    return f"xlsx_{normalized_key(parts[0] if parts else 'station')}_{digest}"


def deterministic_coordinate(city: str, locality: str, index: int) -> tuple[float, float]:
    city_key = city.strip().lower()
    base = CITY_CENTERS.get(city_key)
    if base is None:
        for key, coords in CITY_CENTERS.items():
            if key in city_key:
                base = coords
                break
    if base is None:
        base = CITY_CENTERS["noida"]
    seed = int(hashlib.sha1(f"{city}|{locality}|{index}".encode("utf-8")).hexdigest()[:8], 16)
    lat_offset = ((seed % 2201) - 1100) / 100000
    lon_offset = (((seed // 2201) % 2201) - 1100) / 100000
    return round(base[0] + lat_offset, 6), round(base[1] + lon_offset, 6)


def infer_city(region_or_city: str, location: str = "") -> str:
    source = f"{region_or_city} {location}".lower()
    for hint, city in REGION_HINTS.items():
        if hint in source:
            return city
    return region_or_city.replace("(West)", "").replace("(South)", "").strip() or "Noida"


def parse_power(value: str) -> float | None:
    matches = re.findall(r"(\d+(?:\.\d+)?)\s*kW", value, flags=re.IGNORECASE)
    if not matches:
        return None
    return max(float(item) for item in matches)


def infer_connectors(value: str) -> list[str]:
    text = value.lower()
    connectors: list[str] = []
    if "dc" in text or "fast" in text:
        connectors.append("CCS2")
    if "ac" in text or "type" in text:
        connectors.append("TYPE2")
    if "chademo" in text:
        connectors.append("CHAdeMO")
    return connectors or ["CCS2", "TYPE2"]


def google_directions_url(latitude: float, longitude: float) -> str:
    return f"https://www.google.com/maps/dir/?api=1&destination={latitude:.6f},{longitude:.6f}&travelmode=driving"


def rows_from_sheet(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    for sheet in workbook.worksheets:
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            continue
        headers = [cell_text(value) for value in raw_rows[0]]
        for index, values in enumerate(raw_rows[1:], start=2):
            record = {headers[i]: cell_text(value) for i, value in enumerate(values) if i < len(headers)}
            if not any(record.values()):
                continue
            record["_sheet"] = sheet.title
            record["_row"] = index
            rows.append(record)
    return rows


def station_doc_from_record(record: dict, import_index: int) -> tuple[dict, list[dict]]:
    now = datetime.now(UTC)
    sheet = record["_sheet"]
    if sheet == "Sheet1":
        brand = record.get("Company") or "Imported EV Operator"
        locality = record.get("Area / Locality") or "NCR"
        city = record.get("City") or infer_city(locality)
        landmark = record.get("Type / Landmark") or "EV charging point"
        name = f"{brand} - {locality}"
        address = f"{landmark}, {locality}, {city}, India"
        connector_types = ["CCS2", "TYPE2"]
        max_power_kw = 30.0
    else:
        brand = record.get("Operator / Brand") or "Imported EV Operator"
        locality = record.get("Location Name / Hub") or "NCR"
        city = infer_city(record.get("Region") or "", locality)
        charger_type = record.get("Charger Type (4-Wheeler)") or ""
        name = f"{brand} - {locality}"
        address = f"{locality}, {city}, India"
        connector_types = infer_connectors(charger_type)
        max_power_kw = parse_power(charger_type) or 50.0

    lat, lon = deterministic_coordinate(city, locality, import_index)
    station_id = stable_station_id(sheet, brand, locality, city)
    provider_key = normalized_key(brand)
    is_shubh = "shubh" in f"{brand} {name}".lower()
    station = {
        "station_id": station_id,
        "name": name,
        "brand": brand,
        "operator_name": brand,
        "operator_id": provider_key,
        "address": address,
        "city": city,
        "state": "Delhi NCR",
        "location": {"type": "Point", "coordinates": [lon, lat]},
        "latitude": lat,
        "longitude": lon,
        "google_maps_url": google_directions_url(lat, lon),
        "maps_url_type": "google_directions_coordinate_url",
        "verification_status": "xlsx_import_pending_live_verification",
        "confidence": "demo_geocoded_from_locality",
        "operational_status": "Imported station - live status pending",
        "access_type": "public imported",
        "connector_types": connector_types,
        "max_power_kw": max_power_kw,
        "tariff_summary": "Tariff updates at station",
        "source_import": "subh power.xlsx",
        "source_sheet": sheet,
        "source_row": record["_row"],
        "source_raw": {k: v for k, v in record.items() if not k.startswith("_")},
        "demo_charging_enabled": is_shubh,
        "demo_charger_id": f"SP-DEMO-{import_index:03d}" if is_shubh else None,
        "status_source": "xlsx_import",
        "is_demo": False,
        "environment": get_settings().app_env,
        "created_at": now,
        "updated_at": now,
        "version": 1,
    }
    connectors = []
    for connector_index, connector_type in enumerate(connector_types, start=1):
        connector_id = f"{station_id}-CP{connector_index:02d}"
        connector = {
            "connector_id": connector_id,
            "evse_id": f"{station_id}-EVSE-1",
            "station_id": station_id,
            "connector_type": connector_type,
            "power_kw": max_power_kw,
            "status": "available" if is_shubh else "status_pending",
            "demo_charging_enabled": is_shubh,
            "source_import": "subh power.xlsx",
            "created_at": now,
            "updated_at": now,
        }
        if is_shubh and connector_index == 1:
            connector["demo_charger_id"] = station["demo_charger_id"]
        connectors.append(connector)
    return station, connectors


async def main(db):
    parser = argparse.ArgumentParser(description="Import Shubh Power investor-demo station workbook into MongoDB.")
    parser.add_argument("--path", default=r"C:\Users\shank\Downloads\subh power.xlsx", help="Path to the .xlsx workbook")
    args = parser.parse_args()
    path = Path(args.path)
    if not path.exists():
        raise FileNotFoundError(path)

    source_rows = rows_from_sheet(path)
    await db.stations.delete_many({"source_import": "subh power.xlsx"})
    await db.connectors.delete_many({"source_import": "subh power.xlsx"})
    report = {"source": str(path), "source_rows": len(source_rows), "imported_rows": 0, "connectors_created": 0}
    provider_ops = []
    operator_ops = []
    station_ops = []
    evse_ops = []
    connector_ops = []
    for index, record in enumerate(source_rows, start=1):
        station, connectors = station_doc_from_record(record, index)
        provider_key = normalized_key(station["brand"])
        provider_ops.append(UpdateOne({"provider_key": provider_key}, {"$setOnInsert": {"provider_key": provider_key, "name": station["brand"], "created_at": datetime.now(UTC)}}, upsert=True))
        operator_ops.append(UpdateOne({"operator_id": station["operator_id"]}, {"$setOnInsert": {"operator_id": station["operator_id"], "name": station["brand"], "created_at": datetime.now(UTC)}}, upsert=True))
        station_ops.append(UpdateOne({"station_id": station["station_id"]}, {"$set": station}, upsert=True))
        evse_ops.append(
            UpdateOne(
                {"evse_id": f"{station['station_id']}-EVSE-1"},
                {"$set": {"evse_id": f"{station['station_id']}-EVSE-1", "station_id": station["station_id"], "status": "status_pending", "source_import": "subh power.xlsx", "updated_at": datetime.now(UTC)}},
                upsert=True,
            )
        )
        for connector in connectors:
            connector_ops.append(UpdateOne({"connector_id": connector["connector_id"]}, {"$set": connector}, upsert=True))
            report["connectors_created"] += 1
        report["imported_rows"] += 1
    if provider_ops:
        await db.providers.bulk_write(provider_ops, ordered=False)
    if operator_ops:
        await db.operators.bulk_write(operator_ops, ordered=False)
    if station_ops:
        await db.stations.bulk_write(station_ops, ordered=False)
    if evse_ops:
        await db.evses.bulk_write(evse_ops, ordered=False)
    if connector_ops:
        await db.connectors.bulk_write(connector_ops, ordered=False)
    await db.import_jobs.insert_one({"import_job_id": f"import_{uuid4().hex[:12]}", **report, "created_at": datetime.now(UTC)})
    print(report)


if __name__ == "__main__":
    run(main)
